"""One-off verification harness (not part of the shipped build): pastes the manual test-pack
CSV fixture into a copy of the generated tracker and asks the `formulas` library to actually
evaluate the workbook's formulas, so the Weakness_Panel's RAG flags can be checked
programmatically instead of only by eye. Not wired into CI -- `formulas` is a heavy,
somewhat-partial Excel-formula engine, kept as a dev-only verification aid.
"""
import csv
import os
import shutil

from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(__file__), "..")
SRC = os.path.join(ROOT, "docs", "tracker", "Swing_Gains_Progress_Tracker.xlsx")
DST = os.path.join(ROOT, "docs", "tracker", "_verify_with_fixture.xlsx")
FIXTURE = os.path.join(ROOT, "docs", "tracker", "csv-fixture-3-blocks.csv")

shutil.copyfile(SRC, DST)

wb = load_workbook(DST)
ws = wb["Data_Import"]

with open(FIXTURE, newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    header = next(reader)
    for r, row in enumerate(reader, start=2):
        for c, value in enumerate(row, start=1):
            if value == "":
                continue
            # numeric columns: setNo(6) reps(7) loadKg(8) rpe(9) metricValue(11)
            if c in (6, 7, 8, 9, 11):
                value = float(value)
            ws.cell(row=r, column=c, value=value)

wb.save(DST)
print(f"Wrote {DST}")

import formulas  # noqa: E402

xl_model = formulas.ExcelModel().loads(DST).finish()
solution = xl_model.calculate()


def get(sheet: str, cell: str):
    key = f"'[{os.path.basename(DST).upper()}]{sheet.upper()}'!{cell}"
    for k, v in solution.items():
        if k.upper() == key:
            try:
                return v.value[0, 0]
            except Exception:
                return v.value
    return f"<not found: {key}>"


checks = [
    ("Dashboard", "C3", "current block"),
    ("Weakness_Panel", "B5", "block1 completed status"),
    ("Weakness_Panel", "B6", "block2 completed status"),
    ("Weakness_Panel", "B7", "block3 completed status"),
    ("Weakness_Panel", "E5", "Back squat STALLED? (block 1 row)"),
    ("Weakness_Panel", "E6", "Back squat STALLED? (block 2 row)"),
    ("Weakness_Panel", "H5", "Bench press STALLED? (block 1 row)"),
    ("Weakness_Panel", "H6", "Bench press STALLED? (block 2 row)"),
    ("Weakness_Panel", "J5", "rom_hip avg (block 1 row)"),
    ("Weakness_Panel", "J6", "rom_hip avg (block 2 row)"),
    ("Weakness_Panel", "K5", "ROM STAGNANT? (block 1 row)"),
    ("Weakness_Panel", "K6", "ROM STAGNANT? (block 2 row)"),
    ("Weakness_Panel", "K7", "ROM STAGNANT? (block 3 row)"),
    ("Weakness_Panel", "B16", "frequency-drift: sessions in window"),
    ("Weakness_Panel", "B17", "frequency-drift: rolling avg/week"),
    ("Weakness_Panel", "B18", "frequency-drift: FLAG"),
]
for sheet, cell, label in checks:
    print(f"{label}: {get(sheet, cell)!r}")
