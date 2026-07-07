"""Builds docs/tracker/Swing_Gains_Progress_Tracker.xlsx (spec sec8.2).

An .xlsx engineered for Google Sheets import: conservative formula set (SUMIFS/MAXIFS/
MINIFS/AVERAGEIFS/COUNTIFS/SUMPRODUCT/INDEX/IF), no macros, no Power Query, no dynamic-array
or LAMBDA functions, standard chart types. Every formula here is evaluated by Excel/Sheets on
open, not by this script -- openpyxl only writes formula text. Post-import verification in
Google Sheets is a manual test-pack item (spec sec8.2), not something this script can do
headlessly.

Design notes (flagged, same spirit as the app's own documented spec deviations):

1. The app's CSV export (sec8.1) gives metric rows a BLANK block column -- a MetricLog isn't
   tied to a SessionTemplate/blockId the way a SetLog is. But this tracker needs "ROM/speed by
   block" (sec8.2 Speed_and_ROM). Column O (_EffectiveBlock) on Data_Import forward-fills the
   last non-blank block value down through following metric rows, on the assumption the sheet
   is pasted in the export's own chronological order (append model) and never manually
   reordered. This is the one place the tracker infers something the raw export doesn't state
   outright.
2. "Stalled lift" here reads as MAX(load in block) <= MIN(load in block) for a completed
   block -- i.e. the load never varied at all within that block. This is a simpler, more
   conservative subset of the app's own reading (src/engine/weaknessSignals.ts compares first
   vs peak chronologically) chosen because spreadsheet MAXIFS/MINIFS need no array formulas or
   row-order lookups, while first-vs-last does. Documented here rather than silently matched to
   the in-app engine's exact semantics.
3. All by-block aggregates use a "Block N: *" wildcard match against Data_Import so the sheet
   never has to know an exact block-name string -- SUMIFS/MAXIFS/COUNTIFS/AVERAGEIFS all
   support * wildcards identically in Excel and Sheets.
4. Column N (_FirstOfSession) flags the first row of each (date, sessionType) pair, so
   "sessions this block" / "session-frequency" can count distinct sessions rather than
   set-rows (every session logs multiple exercises = multiple rows).
"""

import os

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SLATE = "34699E"
COPPER = "C77B3C"
WARM_GREY = "D8D4C9"
WHITE = "FFFFFF"
TEXT = "2A2A28"

HEADER_FILL = PatternFill("solid", fgColor=SLATE)
HEADER_FONT = Font(color=WHITE, bold=True)
TITLE_FONT = Font(color=SLATE, bold=True, size=14)
LABEL_FONT = Font(color=TEXT, bold=True)
NOTE_FONT = Font(color="6B6862", italic=True, size=10)

CSV_COLUMNS = [
    "date", "block", "sessionType", "venue", "exercise", "setNo", "reps", "loadKg", "rpe",
    "metricType", "metricValue", "unit", "device"
]
DATA_ROWS = 2000  # pre-filled helper-formula rows on Data_Import; paste new export rows within this range
BLOCKS = list(range(1, 9))  # the programme's 8 blocks (spec sec3.2) -- year-2 cycling is app-side (P3 stub), not tracked here
MAIN_LIFTS = ["Back squat (barbell)", "Bench press (barbell)"]  # example pair; copy a column's formula pattern to track more


def style_header(ws: Worksheet, row: int, last_col: int) -> None:
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def block_wildcard(block_col: str, seq_cell: str) -> str:
    return f'"Block "&{seq_cell}&": *"'


def build_data_import(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Data_Import")
    for i, name in enumerate(CSV_COLUMNS, start=1):
        ws.cell(row=1, column=i, value=name)
    ws.cell(row=1, column=14, value="_FirstOfSession (helper -- do not delete/edit)")
    ws.cell(row=1, column=15, value="_EffectiveBlock (helper -- do not delete/edit)")
    style_header(ws, 1, 15)
    ws.freeze_panes = "A2"

    for r in range(2, DATA_ROWS + 2):
        # First row of a (date, sessionType) pair -- used to count distinct sessions rather
        # than set-rows (each session logs multiple exercises = multiple rows here).
        ws.cell(
            row=r, column=14,
            value=f'=IF($A{r}="","",IF(COUNTIFS($A$2:$A{r},$A{r},$C$2:$C{r},$C{r})=1,1,0))'
        )
        # Forward-fills the last non-blank block value -- see module docstring note 1. Row 2
        # never references the header row (O1) for its fallback, so a metric row pasted before
        # any set-log has ever appeared resolves to "" rather than the header's own text.
        if r == 2:
            ws.cell(row=r, column=15, value=f'=IF($A{r}="","",IF($B{r}<>"",$B{r},""))')
        else:
            ws.cell(row=r, column=15, value=f'=IF($A{r}="","",IF($B{r}<>"",$B{r},$O{r - 1}))')

    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["E"].width = 26

    note = ws.cell(row=1, column=17, value="Paste the app's CSV export starting at row 2 (append model -- keep chronological order).")
    note.font = NOTE_FONT
    return ws


def build_dashboard(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Swing Gains -- Dashboard"
    ws["A1"].font = TITLE_FONT

    # Row count of pasted data -- everything else on this sheet derives from it.
    ws["F1"] = "_RowCount"
    ws["F1"].font = NOTE_FONT
    ws["G1"] = "=COUNTA(Data_Import!$A$2:$A$%d)" % (DATA_ROWS + 1)

    rows = ws.max_row
    cards = [
        ("Current block", '=IF($G$1=0,"-",INDEX(Data_Import!$O$2:$O$%d,$G$1))' % (DATA_ROWS + 1)),
        ("Sessions this block", '=IF($G$1=0,0,SUMPRODUCT((Data_Import!$O$2:$O$%d=$C$3)*(Data_Import!$N$2:$N$%d)))'
         % (DATA_ROWS + 1, DATA_ROWS + 1)),
        ("Latest swing speed (mph)", _last_matching_value("Data_Import!$J$2:$J$%d" % (DATA_ROWS + 1),
                                                           '"swing_speed"', "Data_Import!$K$2:$K$%d" % (DATA_ROWS + 1))),
        ("Latest ROM -- hip switch (grade)", _last_matching_value("Data_Import!$J$2:$J$%d" % (DATA_ROWS + 1),
                                                                   '"rom_hip"', "Data_Import!$K$2:$K$%d" % (DATA_ROWS + 1))),
        ("Latest ROM -- thoracic (deg)", _last_matching_value("Data_Import!$J$2:$J$%d" % (DATA_ROWS + 1),
                                                               '"rom_thoracic"', "Data_Import!$K$2:$K$%d" % (DATA_ROWS + 1))),
        ("Latest ROM -- toe reach (cm)", _last_matching_value("Data_Import!$J$2:$J$%d" % (DATA_ROWS + 1),
                                                               '"rom_reach"', "Data_Import!$K$2:$K$%d" % (DATA_ROWS + 1))),
    ]
    for i, (label, formula) in enumerate(cards, start=3):
        ws.cell(row=i, column=2, value=label).font = LABEL_FONT
        ws.cell(row=i, column=3, value=formula)

    for lift_row, lift in enumerate(MAIN_LIFTS, start=len(cards) + 4):
        ws.cell(row=lift_row, column=2, value=f"Latest load -- {lift} (kg)").font = LABEL_FONT
        ws.cell(row=lift_row, column=3, value=_last_matching_value(
            "Data_Import!$E$2:$E$%d" % (DATA_ROWS + 1), f'"{lift}"', "Data_Import!$H$2:$H$%d" % (DATA_ROWS + 1)
        ))

    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20
    note = ws.cell(row=lift_row + 2, column=2, value="See Lift_Trends / Speed_and_ROM / Weakness_Panel sheets for detail.")
    note.font = NOTE_FONT
    return ws


def _last_matching_value(cond_range: str, cond_value: str, value_range: str) -> str:
    # "Value of the last row where cond_range = cond_value" without array-entry (CSE): forces
    # array evaluation via SUMPRODUCT/MAX, which both Excel and Sheets accept as a plain
    # single-cell formula. Falls back to "-" if there's no match yet.
    return (
        f'=IFERROR(INDEX({value_range},SUMPRODUCT(MAX(({cond_range}={cond_value})'
        f'*(ROW({cond_range})-ROW(INDEX({cond_range},1,1))+1)))),"-")'
    )


def build_lift_trends(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Lift_Trends")
    ws["A1"] = "Load progression per main lift, by block"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=1, column=6, value="Add another lift: copy a column's formula pattern into the next free column.").font = NOTE_FONT

    header_row = 3
    ws.cell(row=header_row, column=1, value="Block")
    for i, lift in enumerate(MAIN_LIFTS, start=2):
        ws.cell(row=header_row, column=i, value=lift)
    style_header(ws, header_row, 1 + len(MAIN_LIFTS))

    for i, seq in enumerate(BLOCKS, start=header_row + 1):
        ws.cell(row=i, column=1, value=seq)
        for j, lift in enumerate(MAIN_LIFTS, start=2):
            col_letter = get_column_letter(j)
            wildcard = block_wildcard("B", f"$A{i}")
            ws.cell(
                row=i, column=j,
                value=(
                    f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard},Data_Import!$E:$E,{col_letter}${header_row})=0,"",'
                    f'MAXIFS(Data_Import!$H:$H,Data_Import!$O:$O,{wildcard},Data_Import!$E:$E,{col_letter}${header_row}))'
                )
            )

    last_row = header_row + len(BLOCKS)
    chart = LineChart()
    chart.title = "Load by block (kg)"
    chart.y_axis.title = "kg"
    chart.x_axis.title = "Block"
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    data = Reference(ws, min_col=1, max_col=1 + len(MAIN_LIFTS), min_row=header_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{last_row + 3}")

    ws.column_dimensions["A"].width = 10
    for j in range(2, 2 + len(MAIN_LIFTS)):
        ws.column_dimensions[get_column_letter(j)].width = 22
    return ws


def build_speed_and_rom(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Speed_and_ROM")
    ws["A1"] = "Swing speed trend and ROM scores, by block"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    columns = [
        ("Block", None),
        ("Swing speed (mph)", "swing_speed"),
        ("ROM -- hip (grade)", "rom_hip"),
        ("ROM -- thoracic (deg)", "rom_thoracic"),
        ("ROM -- reach (cm)", "rom_reach"),
    ]
    for i, (label, _) in enumerate(columns, start=1):
        ws.cell(row=header_row, column=i, value=label)
    style_header(ws, header_row, len(columns))

    for i, seq in enumerate(BLOCKS, start=header_row + 1):
        ws.cell(row=i, column=1, value=seq)
        for j, (_, metric_type) in enumerate(columns, start=1):
            if metric_type is None:
                continue
            wildcard = block_wildcard("O", f"$A{i}")
            ws.cell(
                row=i, column=j,
                value=(
                    f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard},Data_Import!$J:$J,"{metric_type}")=0,"",'
                    f'AVERAGEIFS(Data_Import!$K:$K,Data_Import!$O:$O,{wildcard},Data_Import!$J:$J,"{metric_type}"))'
                )
            )

    last_row = header_row + len(BLOCKS)
    chart = LineChart()
    chart.title = "Swing speed by block"
    chart.y_axis.title = "mph"
    chart.x_axis.title = "Block"
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    data = Reference(ws, min_col=2, max_col=2, min_row=header_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{last_row + 3}")

    rom_chart = LineChart()
    rom_chart.title = "ROM scores by block"
    rom_chart.x_axis.title = "Block"
    rom_data = Reference(ws, min_col=3, max_col=5, min_row=header_row, max_row=last_row)
    rom_chart.add_data(rom_data, titles_from_data=True)
    rom_chart.set_categories(cats)
    ws.add_chart(rom_chart, f"A{last_row + 20}")

    ws.column_dimensions["A"].width = 10
    for col in "BCDE":
        ws.column_dimensions[col].width = 20
    return ws


def build_weakness_panel(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Weakness_Panel")
    ws["A1"] = "Weakness panel -- spec sec8.2 RAG signals"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Completed block = any block with data except the current one (Dashboard!C3). "
        "See tools/build_progress_tracker.py for the exact formula reading of each signal."
    )
    ws["A2"].font = NOTE_FONT

    header_row = 4
    cols = ["Block", "Is completed?"]
    for lift in MAIN_LIFTS:
        cols += [f"{lift}: max", f"{lift}: min", f"{lift}: STALLED?"]
    cols += ["Swing speed avg", "ROM hip avg", "ROM STAGNANT? (vs previous block)"]
    for i, lift in enumerate(MAIN_LIFTS):
        cols += [f"{lift}: SPEED PLATEAU?"]
    for i, c in enumerate(cols, start=1):
        ws.cell(row=header_row, column=i, value=c)
    style_header(ws, header_row, len(cols))

    n_lifts = len(MAIN_LIFTS)
    col_is_completed = 2
    col_lift_start = 3  # each lift takes 3 columns: max, min, stalled
    col_speed = col_lift_start + n_lifts * 3
    col_rom = col_speed + 1
    col_rom_stagnant = col_rom + 1
    col_plateau_start = col_rom_stagnant + 1

    completed_letter = get_column_letter(col_is_completed)
    for i, seq in enumerate(BLOCKS, start=header_row + 1):
        ws.cell(row=i, column=1, value=seq)
        wildcard = block_wildcard("O", f"$A{i}")
        # "Is this the current block" check: compare this row's block number against the
        # sequence number parsed out of Dashboard!C3's "Block N: ..." text.
        ws.cell(
            row=i, column=col_is_completed,
            value=f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard})=0,"no data",IF($A{i}=IFERROR(VALUE(MID(Dashboard!$C$3,7,FIND(":",Dashboard!$C$3)-7)),-1),"in progress","completed"))'
        )

        for li, lift in enumerate(MAIN_LIFTS):
            max_col = col_lift_start + li * 3
            min_col = max_col + 1
            flag_col = max_col + 2
            max_letter = get_column_letter(max_col)
            min_letter = get_column_letter(min_col)
            ws.cell(
                row=i, column=max_col,
                value=f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard},Data_Import!$E:$E,"{lift}")=0,"",MAXIFS(Data_Import!$H:$H,Data_Import!$O:$O,{wildcard},Data_Import!$E:$E,"{lift}"))'
            )
            ws.cell(
                row=i, column=min_col,
                value=f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard},Data_Import!$E:$E,"{lift}")=0,"",MINIFS(Data_Import!$H:$H,Data_Import!$O:$O,{wildcard},Data_Import!$E:$E,"{lift}"))'
            )
            ws.cell(
                row=i, column=flag_col,
                value=(
                    f'=IF(OR({max_letter}{i}="",${completed_letter}{i}<>"completed"),"",'
                    f'IF({max_letter}{i}<={min_letter}{i},"STALLED","ok"))'
                )
            )

        ws.cell(
            row=i, column=col_speed,
            value=f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard},Data_Import!$J:$J,"swing_speed")=0,"",AVERAGEIFS(Data_Import!$K:$K,Data_Import!$O:$O,{wildcard},Data_Import!$J:$J,"swing_speed"))'
        )
        ws.cell(
            row=i, column=col_rom,
            value=f'=IF(COUNTIFS(Data_Import!$O:$O,{wildcard},Data_Import!$J:$J,"rom_hip")=0,"",AVERAGEIFS(Data_Import!$K:$K,Data_Import!$O:$O,{wildcard},Data_Import!$J:$J,"rom_hip"))'
        )
        # "No grade change across the last two transitions" = this completed block's ROM value
        # vs the PRECEDING block's (i-1, the same adjacent-pair reading the plateau flag below
        # uses) -- not two rows back. Only meaningful once a prior row exists, and only
        # flagged when THIS row is itself a completed block.
        if i - 1 >= header_row + 1:
            rom_letter = get_column_letter(col_rom)
            ws.cell(
                row=i, column=col_rom_stagnant,
                value=(
                    f'=IF(OR({rom_letter}{i}="",{rom_letter}{i-1}="",{completed_letter}{i}<>"completed"),"",'
                    f'IF({rom_letter}{i}={rom_letter}{i-1},"STAGNANT","ok"))'
                )
            )
        else:
            ws.cell(row=i, column=col_rom_stagnant, value="")

        for li, lift in enumerate(MAIN_LIFTS):
            max_col = col_lift_start + li * 3
            max_letter = get_column_letter(max_col)
            min_letter = get_column_letter(max_col + 1)
            speed_letter = get_column_letter(col_speed)
            plateau_col = col_plateau_start + li
            if i - 1 >= header_row + 1:
                ws.cell(
                    row=i, column=plateau_col,
                    value=(
                        f'=IF(OR({max_letter}{i}="",{speed_letter}{i}="",{speed_letter}{i-1}="",{completed_letter}{i}<>"completed"),"",'
                        f'IF(AND({max_letter}{i}>{min_letter}{i},{speed_letter}{i}<={speed_letter}{i-1}),"PLATEAU","ok"))'
                    )
                )
            else:
                ws.cell(row=i, column=plateau_col, value="")

    last_row = header_row + len(BLOCKS)

    # Frequency drift (spec "<2/week rolling") needs a date-based window, not a block-based
    # one -- computed directly from Data_Import via the _FirstOfSession helper column,
    # independent of the by-block table above.
    freq_row = last_row + 3
    ws.cell(row=freq_row, column=1, value="Session-frequency drift (<2/week over the last 4 weeks of logged dates)").font = LABEL_FONT
    ws.cell(row=freq_row + 1, column=1, value="Sessions in window").font = NOTE_FONT
    ws.cell(
        row=freq_row + 1, column=2,
        value=(
            f'=SUMPRODUCT((Data_Import!$A$2:$A${DATA_ROWS + 1}<>"")'
            f'*(Data_Import!$A$2:$A${DATA_ROWS + 1}>=MAX(Data_Import!$A$2:$A${DATA_ROWS + 1})-27)'
            f'*(Data_Import!$N$2:$N${DATA_ROWS + 1}=1))'
        )
    )
    ws.cell(row=freq_row + 2, column=1, value="Rolling average /week").font = NOTE_FONT
    ws.cell(row=freq_row + 2, column=2, value=f"=$B${freq_row + 1}/4")
    ws.cell(row=freq_row + 3, column=1, value="FLAG").font = LABEL_FONT
    ws.cell(row=freq_row + 3, column=2, value=f'=IF($B${freq_row + 1}=0,"no data yet",IF($B${freq_row + 2}<2,"DRIFTING","ok"))')

    for col in range(1, col_plateau_start + n_lifts):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 10
    return ws


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    build_data_import(wb)
    build_dashboard(wb)
    build_lift_trends(wb)
    build_speed_and_rom(wb)
    build_weakness_panel(wb)

    wb.active = 1  # Dashboard shown first on open
    out_dir = os.path.join(os.path.dirname(__file__), "..", "docs", "tracker")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Swing_Gains_Progress_Tracker.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
